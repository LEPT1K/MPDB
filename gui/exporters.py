#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Экспорт баз данных и сводных отчётов в форматы CSV, XLSX и PDF
"""

import csv
import io
import json
from datetime import datetime
from pathlib import Path
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.chart import BarChart, Reference

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.legends import Legend

# Отображаемые названия баз данных
DB_TITLES = {
    'capec_database': 'CAPEC',
    'cwe_database': 'CWE',
    'cve_database': 'CVE',
    'mitre_attack': 'MITRE ATT&CK',
}

# Названия баз по ключам графа связей (LinkGraph)
LINK_DB_LABELS = {
    'capec': 'CAPEC',
    'cwe': 'CWE',
    'attack': 'MITRE ATT&CK',
    'cve': 'CVE',
}

# ==================== Шрифты с поддержкой кириллицы для PDF ====================
try:
    _fonts_dir = Path("C:/Windows/Fonts")
    pdfmetrics.registerFont(TTFont("RU", str(_fonts_dir / "arial.ttf")))
    pdfmetrics.registerFont(TTFont("RU-Bold", str(_fonts_dir / "arialbd.ttf")))
    BASE_FONT = "RU"
    BOLD_FONT = "RU-Bold"
except Exception:
    BASE_FONT = "Helvetica"
    BOLD_FONT = "Helvetica-Bold"


# ==================== Общие утилиты ====================

def _flatten(value):
    """Преобразует значение поля записи в строку для табличного представления"""
    if value is None:
        return ""
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _is_empty(value) -> bool:
    if value is None:
        return True
    if isinstance(value, (list, dict, str)):
        return len(value) == 0
    return False


def _collect_columns(data: list) -> list:
    """Собирает объединённый список колонок по всем записям, сохраняя порядок появления"""
    columns = []
    seen = set()
    for record in data:
        for key in record.keys():
            if key not in seen:
                seen.add(key)
                columns.append(key)
    return columns


# ==================== CSV ====================

def export_csv(data: list) -> io.BytesIO:
    """Экспортирует записи базы данных в CSV (UTF-8 с BOM, разделитель ';' для Excel)"""
    columns = _collect_columns(data)

    text_buffer = io.StringIO()
    writer = csv.writer(text_buffer, delimiter=';', lineterminator='\r\n')
    writer.writerow(columns)
    for record in data:
        writer.writerow([_flatten(record.get(col)) for col in columns])

    output = io.BytesIO()
    output.write(b'\xef\xbb\xbf')  # BOM, чтобы Excel корректно определил UTF-8
    output.write(text_buffer.getvalue().encode('utf-8'))
    output.seek(0)
    return output


# ==================== XLSX ====================

def export_xlsx(data: list, db_title: str) -> io.BytesIO:
    """Экспортирует записи базы данных в XLSX: лист с данными + лист сводки с диаграммой заполненности полей"""
    columns = _collect_columns(data)

    wb = Workbook()
    ws = wb.active
    ws.title = "Данные"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    ws.append(columns)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for record in data:
        ws.append([_flatten(record.get(col)) for col in columns])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment

    sample_rows = min(ws.max_row, 50)
    for i, col in enumerate(columns, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        max_len = len(col)
        for r in range(2, sample_rows + 1):
            value = ws.cell(row=r, column=i).value
            if value:
                max_len = max(max_len, min(len(str(value)), 60))
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 60)

    ws.freeze_panes = "A2"

    # Лист сводки
    summary = wb.create_sheet("Сводка")
    summary["A1"] = f"Сводка по базе: {db_title}"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A3"] = "Всего записей:"
    summary["B3"] = len(data)
    summary["A4"] = "Дата экспорта:"
    summary["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    header_row = 6
    summary.cell(row=header_row, column=1, value="Поле")
    summary.cell(row=header_row, column=2, value="Заполнено")
    summary.cell(row=header_row, column=3, value="Пусто")
    for col_idx in range(1, 4):
        cell = summary.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    row = header_row + 1
    for col in columns:
        filled = sum(1 for record in data if not _is_empty(record.get(col)))
        empty = len(data) - filled
        summary.cell(row=row, column=1, value=col)
        summary.cell(row=row, column=2, value=filled)
        summary.cell(row=row, column=3, value=empty)
        row += 1
    last_row = row - 1

    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 12
    summary.column_dimensions["C"].width = 12

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.style = 10
    chart.title = "Заполненность полей"
    chart.y_axis.title = "Записей"
    data_ref = Reference(summary, min_col=2, max_col=3, min_row=header_row, max_row=last_row)
    cats_ref = Reference(summary, min_col=1, min_row=header_row + 1, max_row=last_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 11
    chart.width = 22
    summary.add_chart(chart, "E6")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ==================== PDF-отчёт ====================

def _bar_chart_drawing(categories, values, bar_color, width=440, height=200):
    """Рисует столбчатую диаграмму для вставки в PDF"""
    drawing = Drawing(width, height)
    chart = VerticalBarChart()
    chart.x = 50
    chart.y = 50
    chart.height = height - 70
    chart.width = width - 80
    chart.data = [values]
    chart.categoryAxis.categoryNames = categories
    chart.categoryAxis.labels.boxAnchor = 'ne'
    chart.categoryAxis.labels.angle = 20
    chart.categoryAxis.labels.dx = -5
    chart.categoryAxis.labels.fontName = BASE_FONT
    chart.valueAxis.labels.fontName = BASE_FONT
    chart.valueAxis.valueMin = 0
    chart.bars[0].fillColor = bar_color
    chart.barLabels.fontName = BASE_FONT
    chart.barLabelFormat = '%d'
    chart.barLabels.nudge = 10
    drawing.add(chart)
    return drawing


def _pie_chart_drawing(labels, values, width=440, height=220):
    """Рисует круговую диаграмму распределения записей по базам для вставки в PDF"""
    drawing = Drawing(width, height)
    chart = Pie()
    chart.x = 70
    chart.y = 20
    chart.width = 160
    chart.height = 160
    chart.data = values
    chart.labels = [f"{label} ({value})" for label, value in zip(labels, values)]
    chart.simpleLabels = False
    chart.slices.strokeWidth = 0.5

    palette = [colors.HexColor('#0d6efd'), colors.HexColor('#198754'),
               colors.HexColor('#ffc107'), colors.HexColor('#dc3545'),
               colors.HexColor('#6f42c1'), colors.HexColor('#fd7e14')]
    for i in range(len(values)):
        chart.slices[i].fillColor = palette[i % len(palette)]
        chart.slices[i].fontName = BASE_FONT
        chart.slices[i].fontSize = 8

    drawing.add(chart)

    legend = Legend()
    legend.x = 280
    legend.y = height - 30
    legend.dx = 8
    legend.dy = 8
    legend.fontName = BASE_FONT
    legend.fontSize = 8
    legend.alignment = 'left'
    legend.columnMaximum = len(values)
    legend.colorNamePairs = [
        (palette[i % len(palette)], f"{label}: {value}")
        for i, (label, value) in enumerate(zip(labels, values))
    ]
    drawing.add(legend)
    return drawing


def _styled_table(table_data, col_widths, bold_last_row=True):
    """Таблица в едином стиле отчёта"""
    table = Table(table_data, colWidths=col_widths)
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), BASE_FONT),
        ('FONTNAME', (0, 0), (-1, 0), BOLD_FONT),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]
    if bold_last_row:
        style += [
            ('FONTNAME', (0, -1), (-1, -1), BOLD_FONT),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f1f3f5')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e9ecef')),
        ]
    else:
        style.append(('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f3f5')]))
    table.setStyle(TableStyle(style))
    return table


def export_pdf_report(databases: list, link_stats: dict = None) -> io.BytesIO:
    """Формирует сводный PDF-отчёт: статистика баз, расшифровка перекрёстных связей и графики.

    databases — список словарей из /api/databases (name, records, links);
    link_stats — статистика LinkGraph.get_link_statistics() для расшифровки связей.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title="Отчёт MPDB"
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleRu", parent=styles["Title"], fontName=BOLD_FONT)
    heading_style = ParagraphStyle("HeadingRu", parent=styles["Heading2"], fontName=BOLD_FONT)
    normal_style = ParagraphStyle("NormalRu", parent=styles["Normal"], fontName=BASE_FONT)
    note_style = ParagraphStyle("NoteRu", parent=styles["Normal"], fontName=BASE_FONT,
                                fontSize=9, textColor=colors.HexColor('#495057'))

    elements = []
    elements.append(Paragraph("Отчёт по базам данных MPDB", title_style))
    elements.append(Paragraph(
        f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}", normal_style))
    elements.append(Spacer(1, 0.4 * cm))
    elements.append(Paragraph(
        "MPDB (MITRE Parser Databases) объединяет четыре базы в области кибербезопасности: "
        "CAPEC (шаблоны атак), CWE (типы уязвимостей), CVE (известные уязвимости, NVD) и "
        "MITRE ATT&amp;CK (тактики и техники злоумышленников). Отчёт построен по локальным данным, "
        "полученным из официальных источников на шагах «Парсинг», «Связывание» и «Автозаполнение».",
        normal_style))
    elements.append(Spacer(1, 0.8 * cm))

    names = [db.get('name', db.get('id', '')) for db in databases]
    records = [db.get('records', 0) for db in databases]
    links = [db.get('links', 0) for db in databases]

    # Сводная таблица: записи и исходящие связи каждой базы
    table_data = [["База данных", "Записей", "Перекрёстных связей"]]
    for name, rec, link in zip(names, records, links):
        table_data.append([name, str(rec), str(link)])
    table_data.append(["Итого", str(sum(records)), str(sum(links))])

    elements.append(Paragraph("Общая статистика", heading_style))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(_styled_table(table_data, [6 * cm, 3.5 * cm, 5 * cm]))
    elements.append(Spacer(1, 0.4 * cm))
    elements.append(Paragraph(
        "<b>Запись</b> — один элемент базы: шаблон атаки (CAPEC), тип уязвимости (CWE), "
        "конкретная уязвимость (CVE) или техника атакующих (ATT&amp;CK). Количество записей "
        "ограничено лимитами парсинга в настройках, поэтому может быть меньше полного объёма "
        "официальной базы.", note_style))
    elements.append(Spacer(1, 0.2 * cm))
    elements.append(Paragraph(
        "<b>Перекрёстная связь</b> — одна заполненная ссылка из записи на запись другой базы "
        "в полях related_capec, related_cwe, related_mitre и related_cve. Связи устанавливаются "
        "процессами «Связывание» (по данным первоисточников) и «Автозаполнение» (восстановление "
        "обратных и транзитивных ссылок). Число связей базы — это сумма таких ссылок по всем её "
        "записям, поэтому оно может многократно превышать число записей.", note_style))

    # Пример расшифровки на базе с наибольшим числом связей
    if links and max(links) > 0:
        idx = links.index(max(links))
        if records[idx] > 0:
            elements.append(Spacer(1, 0.2 * cm))
            elements.append(Paragraph(
                f"Например, {links[idx]} связей базы {names[idx]} — это все ссылки из её "
                f"{records[idx]} записей на записи других баз, в среднем "
                f"{links[idx] / records[idx]:.1f} ссылки на запись.", note_style))
    elements.append(Spacer(1, 0.8 * cm))

    # Расшифровка: из чего складываются связи (по направлениям между базами)
    edges = (link_stats or {}).get('edges', [])
    if edges:
        elements.append(Paragraph("Из чего складываются связи", heading_style))
        elements.append(Spacer(1, 0.3 * cm))
        elements.append(Paragraph(
            "Каждая строка — количество ссылок из записей одной базы на записи другой "
            "(и поле, в котором они хранятся). Сумма по таблице равна общему числу связей.",
            note_style))
        elements.append(Spacer(1, 0.3 * cm))

        edges_sorted = sorted(edges, key=lambda e: e.get('count', 0), reverse=True)
        edge_table = [["Откуда", "Куда", "Поле", "Связей"]]
        for edge in edges_sorted:
            edge_table.append([
                LINK_DB_LABELS.get(edge.get('source'), edge.get('source', '')),
                LINK_DB_LABELS.get(edge.get('target'), edge.get('target', '')),
                edge.get('field', ''),
                str(edge.get('count', 0)),
            ])
        edge_table.append(["Итого", "", "", str(sum(e.get('count', 0) for e in edges_sorted))])
        elements.append(_styled_table(edge_table, [4 * cm, 4 * cm, 4 * cm, 2.5 * cm]))
        elements.append(Spacer(1, 0.8 * cm))

    # Диаграмма: записи по базам
    elements.append(Paragraph("Количество записей по базам", heading_style))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(_bar_chart_drawing(names, records, colors.HexColor('#0d6efd')))
    elements.append(Spacer(1, 1 * cm))

    # Диаграмма: связи по базам
    elements.append(Paragraph("Количество перекрёстных связей по базам", heading_style))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(_bar_chart_drawing(names, links, colors.HexColor('#198754')))
    elements.append(Spacer(1, 1 * cm))

    # Круговая диаграмма: доля записей по базам
    if sum(records) > 0:
        elements.append(Paragraph("Распределение записей по базам", heading_style))
        elements.append(Spacer(1, 0.3 * cm))
        elements.append(_pie_chart_drawing(names, records))
        elements.append(Spacer(1, 0.8 * cm))

    # Самые цитируемые записи (на них чаще всего ссылаются другие базы)
    top_referenced = (link_stats or {}).get('top_referenced', [])
    if top_referenced:
        elements.append(Paragraph("Самые цитируемые записи", heading_style))
        elements.append(Spacer(1, 0.3 * cm))
        elements.append(Paragraph(
            "Записи, на которые чаще всего ссылаются записи других баз (по числу входящих связей). "
            "Подробный граф связей каждой записи доступен на странице «Связи» веб-интерфейса.",
            note_style))
        elements.append(Spacer(1, 0.3 * cm))

        top_table = [["ID", "Название", "База", "Входящих ссылок"]]
        for item in top_referenced:
            name = item.get('name') or ''
            if len(name) > 60:
                name = name[:57] + '...'
            top_table.append([
                item.get('id', ''),
                Paragraph(xml_escape(name), note_style),
                LINK_DB_LABELS.get(item.get('db'), item.get('db', '')),
                str(item.get('count', 0)),
            ])
        elements.append(_styled_table(top_table, [3 * cm, 7 * cm, 3 * cm, 3 * cm],
                                      bold_last_row=False))

    doc.build(elements)
    buffer.seek(0)
    return buffer
